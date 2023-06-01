# Web service de Ley Chile para desarrolladores
# https://www.bcn.cl/leychile/consulta/legislacion_abierta_web_service
#

import openpyxl
from openpyxl.styles import Font

import requests
#from bs4 import BeautifulSoup
#import json
import time
import random
import xml.etree.ElementTree as ET
import datetime

# Cargar el archivo de Excel con las URLs de la normativa
archivo_excel = openpyxl.load_workbook('marco_normativo.xlsx')
#hoja_excel=sheet = workbook['chao']
hoja_excel = archivo_excel.active
ultima_columna=hoja_excel.max_column
normaBCN_meta=True #si es falso buscará la URL del XML de la norma full en vez de la URL que solo tienen metadadata de la norma


# Recorrer las filas de la hoja de Excel
for indice, fila in enumerate(hoja_excel.iter_rows(min_row=80, max_row=80,min_col=8, max_col=8, values_only=True), start=4): #max_row=80,
    # Obtener la dirección web que está en la columna F
    url = fila[0]
    #print("URL= ",url)
    if url is not None and (url.find("bcn.cl") != -1 or url.find("leychile.cl") != -1):
             # Obtener el id de la norma desde la URL
            id_norma = None
            if "idNorma=" in url:
                id_norma = url.split("idNorma=")[1].split("&")[0]
           
           
           #para avisar posterirmente que se está usando una versión de fecha fija en vez de una url genérica.
            if "idVersion=" in url:
                id_version = url.split("idVersion=")[1].split("&")[0]
            else :
                id_version=False




            if id_norma is not None:

                if normaBCN_meta:
                    url2 = f"https://www.leychile.cl/Consulta/obtxml?opt=4546&idNorma={id_norma}"
                else :
                    url2 = f"https://servicios-leychile.bcn.cl/Consulta/obtxml?opt=7&idNorma={id_norma}"
                # Construir la URL para obtener los datos de la norma en base a un XML de la norma completa (Última versión de cualquier norma (leyes, decretos, tratados, DL, DFL, resoluciones, etc.))
                # https://www.bcn.cl/leychile/consulta/legislacion_abierta_web_service
                #url2 = f"https://servicios-leychile.bcn.cl/Consulta/obtxml?opt=7&idNorma={id_norma}"
                #url2=f"https://www.leychile.cl/Consulta/obtxml?opt=4546&idNorma=1146551" #URL DE EJEMPLO (para debug)
                

                #URL SOLO DE METDATA de la NOrma
                #url2 = f"https://www.leychile.cl/Consulta/obtxml?opt=4546&idNorma={id_norma}"
                #url2 = f"https://www.leychile.cl/Consulta/obtxml?opt=4546&idNorma=1146551"



                #time.sleep(random.uniform(0.1, 0.5))
                headers = { 
                            "accept": "application/xml"
                        }
                #print("url2 :",url2,"\n\n\n")
                respuesta = requests.get(url2, headers=headers)
                xml_data = respuesta.content
                arbol = ET.ElementTree(ET.fromstring(xml_data))
                raiz = arbol.getroot()
                
                """  
                #INCIO Sección para resctar campos asumiendo que el XML es de la norma completa (f"https://servicios-leychile.bcn.cl/Consulta/obtxml?opt=7&idNorma={id_norma}")
                #Última versión de cualquier norma (leyes, decretos, tratados, DL, DFL, resoluciones, etc.)
                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Identificador') #Promulgacion -- Publicacion
                fechas_norma = tmp.attrib
                
                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Tipo') #Tipo Norma
                tipo_norma = tmp.text


                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Numero') #Numero norma
                numero_norma = tmp.text 

                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}TituloNorma') #Titulo Norma
                titulo_norma = tmp.text 

                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Norma') #Vigencia norma
                norma_derogada=raiz.attrib.get("derogado")
                nombre2 = raiz.find('Identificador')

                print("Norma Orgánica,",tipo_norma,"|",numero_norma,"|",titulo_norma,"|",datetime.datetime.strptime(fechas_norma.get("fechaPublicacion"), "%Y-%m-%d").strftime("%d/%m/%Y"),"|",url,",", datetime.datetime.strptime(fechas_norma.get("fechaPromulgacion"), "%Y-%m-%d").strftime("%d/%m/%Y"))
                #FIN de Sección para resctar campos asumiendo que el XML es de la norma completa 
                """
                
                #
                #INCIO Sección para resctar campos asumiendo que el XML es de la norma Consulta por los metadatos de una norma (Última versión de cualquier norma (leyes, decretos, tratados, DL, DFL, resoluciones, etc.))
                # Este es la URL genrica  (f"https://servicios-leychile.bcn.cl/Consulta/obtxml?opt=7&idNorma={id_norma}")
                #
                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Identificador') #Fecha Publicacion
                fecha_publicacion = tmp.attrib.get("fechaPublicacion")
                fecha_publicacion = datetime.datetime.strptime(fecha_publicacion, "%Y-%m-%d").strftime("%d/%m/%Y") #Formatea a DD/MM/YYYY

                #fecha_promulgacion=tmp.attrib.get("fechaPromulgacion")
                #fecha_promulgacion=datetime.datetime.strptime(fecha_promulgacion, "%Y-%m-%d").strftime("%d/%m/%Y") #Formatea a DD/MM/YYYY
                
                caracteres_a_eliminar = ['\n', '\r', '\b', '\t']
                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Encabezado') #Fecha ultima versión y su vigencia
                ultima_version=tmp.attrib.get("fechaVersion")
                ultima_version = datetime.datetime.strptime(ultima_version, "%Y-%m-%d").strftime("%d/%m/%Y") #Formatea a DD/MM/YYYY
                vigencia=tmp.attrib.get("derogado")

  

                tmp_list = raiz.findall('.//{http://www.leychile.cl/esquemas}Tipo')  # Tipos de norma
                tipo_norma = tmp_list[0].text.strip() if tmp_list else ""
                for caracter in caracteres_a_eliminar:
                    tipo_norma = tipo_norma.replace(caracter, '')

                tmp_list = raiz.findall('.//{http://www.leychile.cl/esquemas}Numero')  # Números de norma
                if tmp_list:
                    numero_norma = tmp_list[-1].text.strip()
                else:
                    numero_norma = ""
                for caracter in caracteres_a_eliminar :
                    numero_norma = numero_norma.replace(caracter, '')
                solo_digitos = "".join(filter(str.isdigit, numero_norma))
                sin_digitos = "".join(filter(lambda x: not x.isdigit(), numero_norma))
                tipo_norma = tipo_norma.title() + sin_digitos.title()  # Agrega la palabra EXENTA o EXENTO al tipo de norma
                for caracter in ["\\","-","/"]:
                     tipo_norma = tipo_norma.replace(caracter, '')
                for caracter in [ "EXENTO","Exento","EXENTA","Exenta"]:
                     numero_norma = numero_norma.replace(caracter, '')

                """ ESTA SECCIÓN Solo busca la primera aparicion del numero de la NORMA (hay problemas cuando tiene varios nuemeros ej: https://www.bcn.cl/leychile/navegar?idNorma=243771)
                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Tipo') #Tipo Norma
                tipo_norma = tmp.text.strip()
                for caracter in caracteres_a_eliminar:
                    tipo_norma = tipo_norma.replace(caracter, '')

                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Numero') #Numero norma
                numero_norma = tmp.text.strip()
                for caracter in caracteres_a_eliminar:
                    numero_norma = numero_norma.replace(caracter, '')
                solo_digitos = ""
                sin_digitos = ""
                for caracter in numero_norma:
                    if caracter.isdigit():
                        solo_digitos += caracter
                    else:
                        sin_digitos += caracter
                tipo_norma = tipo_norma.title() + sin_digitos.title()  #agraga la palabra EXENTA o EXENTO al tipo de norma que puede ser  "Resolución o Decreto"
                numero_norma=solo_digitos
                """

                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}TituloNorma') #Titulo Norma
                titulo_norma = tmp.text.strip()
                for caracter in caracteres_a_eliminar:
                    titulo_norma = titulo_norma.replace(caracter, '')


                tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Organismo') #Nombre del Organizmo
                organismo = tmp.text.strip()
                for caracter in caracteres_a_eliminar:
                    organismo = organismo.replace(caracter, '')


                norma_derogada=raiz.attrib.get("derogado")
                nombre2 = raiz.find('Identificador')

               ### Escribe en las columnas del archivo excell
                siguiente_columna=ultima_columna + 1
                hoja_excel.cell(row=indice, column=siguiente_columna, value=organismo)
                siguiente_columna +=1

                hoja_excel.cell(row=indice, column=siguiente_columna, value=tipo_norma)
                siguiente_columna +=1

                hoja_excel.cell(row=indice, column=siguiente_columna, value=numero_norma)
                siguiente_columna +=1

                hoja_excel.cell(row=indice, column=siguiente_columna, value=titulo_norma)
                siguiente_columna +=1

                hoja_excel.cell(row=indice, column=siguiente_columna, value=fecha_publicacion)
                siguiente_columna +=1

                hoja_excel.cell(row=indice, column=siguiente_columna, value=ultima_version)
                siguiente_columna +=1
                
                hoja_excel.cell(row=indice, column=siguiente_columna, value=url)
                siguiente_columna +=1


                if "no derogado" not in vigencia:
                    hoja_excel.cell(row=indice, column=siguiente_columna).font = Font(color="FF0000")  # cambia color rojo si es distinto a "no derogado"
                elif id_version:
                    hoja_excel.cell(row=indice, column=siguiente_columna).font = Font(color="FF0000")  # cambia color rojo Id_version existe en el URL orginal del archivo excel (lo que quiere decir que está revisando una fecha fija no la ultima)
                    vigencia = vigencia+" "+"(la URL original apunta a una versión FIJA de "+id_version+" )"
                hoja_excel.cell(row=indice, column=siguiente_columna, value=vigencia)
                

                siguiente_columna =ultima_columna #Como es el final del ciclo se re inicia la poscion
                
                print(organismo,"|",tipo_norma,"|",numero_norma,"|",titulo_norma,"|",fecha_publicacion,"|",ultima_version,"|",url,"|",vigencia)


                #
                #
                #FIN de Sección para resctar campos asumiendo que el XML es de la norma es solo metadata
                #
                #
    else:
        hoja_excel.cell(row=indice, column=ultima_columna+1).font = Font(color="FF0000")  # cambia color rojo Id_version existe en el URL orginal del archivo excel (lo que quiere decir que está revisando una fecha fija no la ultima)
        hoja_excel.cell(row=indice, column=ultima_columna+1, value="No es una URL de BCN.cl o LeyChile.cl")           
archivo_excel.save(filename='marco_normativo.xlsx')
