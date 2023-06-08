# Web service de Ley Chile para desarrolladores
# https://www.bcn.cl/leychile/consulta/legislacion_abierta_web_service
#

from openpyxl.styles import Font

import requests

import xml.etree.ElementTree as ET
import datetime
import os
from queue import Queue
import time


def recorre_filas_excel(archivo_excel,esMetaData,cliente_ip,nombre_archivo,cola):
    #hoja_excel=sheet = workbook['chao']
    hoja_excel = archivo_excel.active
    ultima_columna=hoja_excel.max_column
    # Recorrer las filas de la hoja de Excel
    for indice, fila in enumerate(hoja_excel.iter_rows(min_row=4, min_col=8, max_col=8, values_only=True), start=4): #max_row=80,
       
        # Obtener la dirección web que está en la columna F
        url = fila[0]
        #print("URL= ",url)
        if url is not None and (url.find("bcn.cl") != -1 or url.find("leychile.cl") != -1):
                # Obtener el id de la norma desde la URL
                id_norma = None
                if "idNorma=" in url:
                    id_norma = url.split("idNorma=")[1].split("&")[0]
            
            
            #para avisar posterirmente que se está usando una versión de fecha fija en vez de una url genérica.
                if "idVersion=" in url:
                    id_version = url.split("idVersion=")[1].split("&")[0]
                else :
                    id_version=False

                url="https://www.bcn.cl/leychile/navegar?idNorma="+id_norma


                if id_norma is not None:

                    if esMetaData:
                        datos_norma=busca_datos(esMetaData,id_norma,False)  ##WTF!!! xd
                    else :
                        datos_norma=busca_datos(esMetaData,id_norma,False) ##WTF!!! xd
                
                ### Escribe en las columnas del archivo excell
                    siguiente_columna=ultima_columna + 1
                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["organismo"])
                    siguiente_columna +=1

                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["tipo_norma"])
                    siguiente_columna +=1

                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["numero_norma"])
                    siguiente_columna +=1

                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["titulo_norma"])
                    siguiente_columna +=1

                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["fecha_promulgacion"])
                    siguiente_columna +=1
                    
                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["fecha_publicacion"])
                    siguiente_columna +=1

                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["ultima_version"])
                    siguiente_columna +=1
                    
                    hoja_excel.cell(row=indice, column=siguiente_columna, value=url)
                    siguiente_columna +=1

                    if "no derogado" not in datos_norma["vigencia"]:
                        hoja_excel.cell(row=indice, column=siguiente_columna).font = Font(color="FF0000")  # cambia color rojo si es distinto a "no derogado"
                    elif id_version:
                        hoja_excel.cell(row=indice, column=siguiente_columna).font = Font(color="FF0000")  # cambia color rojo Id_version existe en el URL orginal del archivo excel (lo que quiere decir que está revisando una fecha fija no la ultima)
                        datos_norma["vigencia"] =datos_norma["vigencia"]+" "+"(la URL original apunta a una versión FIJA de "+id_version+" )"

                    hoja_excel.cell(row=indice, column=siguiente_columna, value=datos_norma["vigencia"])
                    

                    siguiente_columna =ultima_columna #Como es el final del ciclo se re inicia la poscion
                    
                    #print(datos_norma["organismo"],"|",datos_norma["tipo_norma"],"|",datos_norma["numero_norma"],"|",datos_norma["titulo_norma"],"|",datos_norma["fecha_promulgacion"],"|",datos_norma["fecha_publicacion"],"|",datos_norma["ultima_version"],"|",url,"|",datos_norma["vigencia"])
                    print(datos_norma)

                    cola.put(datos_norma) 

        else:
            hoja_excel.cell(row=indice, column=ultima_columna+1).font = Font(color="FF0000")  # cambia color rojo Id_version existe en el URL orginal del archivo excel (lo que quiere decir que está revisando una fecha fija no la ultima)
            hoja_excel.cell(row=indice, column=ultima_columna+1, value="No es una URL de BCN.cl o LeyChile.cl")
    
    print ("_FIN_DATOS_")
    filename="ok"+nombre_archivo        
    archivo_excel.save(filename)
    return filename  #Regresa el nombre del archivo excel procesado



def busca_datos(esMetaData,id_norma,esRecurisivo=False):

    if esMetaData:
        url2=f"https://www.leychile.cl/Consulta/obtxml?opt=4546&idNorma={id_norma}"
    else:
        url2 =f"https://servicios-leychile.bcn.cl/Consulta/obtxml?opt=7&idNorma={id_norma}"


    #time.sleep(random.uniform(0.1, 0.5))
    headers = { 
                "accept": "application/xml"
            }    
    respuesta = requests.get(url2, headers=headers)
    xml_data = respuesta.content
    #arbol = ET.ElementTree(ET.fromstring(xml_data))


    try:  #Si hay un error en la respuesta de BCN, pruena con el otro ULR y si vuelve a falla, devuele nulos
        arbol = ET.ElementTree(ET.fromstring(xml_data))
    except :
        if esRecurisivo:  
            return {
                "fecha_promulgacion":"",
                "fecha_publicacion": "",
                "ultima_version": "",
                "vigencia": "",
                "tipo_norma": "",
                "numero_norma": "",
                "titulo_norma": "",
                "organismo": "ERROR EN CONSULTAR BNC: "+url2
            }
        else:
                return (busca_datos(not esMetaData,id_norma,True))
           
    raiz = arbol.getroot()

    caracteres_a_eliminar = ['\n', '\r', '\b', '\t']

    ultima_version = raiz.attrib.get("fechaVersion") #Ultima Version
    ultima_version = datetime.datetime.strptime(ultima_version, "%Y-%m-%d").strftime("%d/%m/%Y") #Formatea a DD/MM/YYYY


    tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Encabezado') #vigencia
    vigencia=tmp.attrib.get("derogado")


    tmp_list = raiz.findall('.//{http://www.leychile.cl/esquemas}Tipo')  # Tipos de norma
    tipo_norma = tmp_list[0].text.strip() if tmp_list else ""
    for caracter in caracteres_a_eliminar:
        tipo_norma = tipo_norma.replace(caracter, '')

    tmp_list = raiz.findall('.//{http://www.leychile.cl/esquemas}Numero')  # Números de norma
    if tmp_list:
        numero_norma = tmp_list[-1].text.strip()
    else:
        numero_norma = ""
    for caracter in caracteres_a_eliminar :
        numero_norma = numero_norma.replace(caracter, '')
    solo_digitos = "".join(filter(str.isdigit, numero_norma))
    sin_digitos = "".join(filter(lambda x: not x.isdigit(), numero_norma))
    tipo_norma = tipo_norma.title() + sin_digitos.title()  # Agrega la palabra EXENTA o EXENTO al tipo de norma
    for caracter in ["\\","-","/"]:
            tipo_norma = tipo_norma.replace(caracter, '')
    for caracter in [ "EXENTO","Exento","EXENTA","Exenta"]:
            numero_norma = numero_norma.replace(caracter, '')

    tmp = raiz.find('.//{http://www.leychile.cl/esquemas}TituloNorma') #Titulo Norma
    titulo_norma = tmp.text.strip()
    for caracter in caracteres_a_eliminar:
        titulo_norma = titulo_norma.replace(caracter, '')


    tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Organismo') #Nombre del Organizmo
    organismo = tmp.text.strip()
    for caracter in caracteres_a_eliminar:
        organismo = organismo.replace(caracter, '')



    if esMetaData:
        tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Identificador') #Fecha Publicacion
        fecha_publicacion = tmp.attrib.get("fechaPublicacion")
        fecha_publicacion = datetime.datetime.strptime(fecha_publicacion, "%Y-%m-%d").strftime("%d/%m/%Y") #Formatea a DD/MM/YYYY

        fecha_promulgacion="Sin información"
 
    else:
        tmp = raiz.find('.//{http://www.leychile.cl/esquemas}Identificador') #Promulgacion -- Publicacion
        fechas_norma = tmp.attrib
        fecha_promulgacion = fechas_norma["fechaPromulgacion"]
        fecha_promulgacion = datetime.datetime.strptime(fecha_promulgacion, "%Y-%m-%d").strftime("%d/%m/%Y") #Formatea a DD/MM/YYYY  
             
        fecha_publicacion = fechas_norma["fechaPublicacion"]
        fecha_publicacion = datetime.datetime.strptime(fecha_publicacion, "%Y-%m-%d").strftime("%d/%m/%Y") #Formatea a DD/MM/YYYY  

    return {
        "fecha_promulgacion":fecha_promulgacion,
        "fecha_publicacion": fecha_publicacion,
        "ultima_version": ultima_version,
        "vigencia": vigencia,
        "tipo_norma": tipo_norma,
        "numero_norma": numero_norma,
        "titulo_norma": titulo_norma,
        "organismo": organismo
        
    }
